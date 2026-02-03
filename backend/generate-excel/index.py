import json
import io
import base64
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import urllib.request
import urllib.parse
from PIL import Image as PILImage
import psycopg2

def get_next_kp_number():
    """Получить следующий номер КП из счетчика с автосбросом в начале года"""
    try:
        current_year = datetime.now().year
        database_url = os.environ.get('DATABASE_URL')
        
        conn = psycopg2.connect(database_url)
        cur = conn.cursor()
        
        # Получаем текущий счетчик
        cur.execute("SELECT year, counter FROM kp_counter WHERE id = 1")
        row = cur.fetchone()
        
        if row:
            saved_year, counter = row
            # Если год изменился, сбрасываем счетчик
            if saved_year != current_year:
                counter = 0
        else:
            counter = 0
        
        counter += 1
        
        # Обновляем счетчик в базе
        cur.execute(
            "UPDATE kp_counter SET year = %s, counter = %s, updated_at = CURRENT_TIMESTAMP WHERE id = 1",
            (current_year, counter)
        )
        conn.commit()
        
        cur.close()
        conn.close()
        
        return counter
    except Exception as e:
        print(f'Error getting KP number: {e}')
        return 1

def handler(event, context):
    """Генерация Excel или PDF файла с коммерческим предложением"""
    print(f'Function started. Memory limit: {context.memory_limit_in_mb} MB')
    
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': ''
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        products = body.get('products', [])
        print(f'Processing {len(products)} products')
        address = body.get('address', '')
        installation_percent = body.get('installationPercent', 0)
        installation_cost = body.get('installationCost', 0)
        delivery_cost = body.get('deliveryCost', 0)
        hide_installation = body.get('hideInstallation', False)
        hide_delivery = body.get('hideDelivery', False)
        file_format = body.get('format', 'xlsx')
        discount_percent = body.get('discountPercent', 0)
        discount_amount = body.get('discountAmount', 0)
        
        print(f'Discount percent: {discount_percent}')
        print(f'Discount amount: {discount_amount}')
        
        # Получаем номер КП
        kp_number = get_next_kp_number()
        
        # Генерация PDF
        if file_format == 'pdf':
            from pdf_reportlab import generate_pdf_reportlab
            
            pdf_content = generate_pdf_reportlab(
                products, address, installation_percent, installation_cost,
                delivery_cost, hide_installation, hide_delivery, kp_number,
                discount_percent, discount_amount
            )
            
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': 'attachment; filename="commercial_offer.pdf"',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': base64.b64encode(pdf_content).decode('utf-8'),
                'isBase64Encoded': True
            }
        
        # Генерация Excel (по умолчанию)
        wb = Workbook()
        ws = wb.active
        ws.title = "Коммерческое предложение"
        
        # Убираем отступы сверху
        ws.page_margins.top = 0
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        
        # Отступ одной строки сверху
        current_row = 2
        
        # Логотип (левый верхний угол) - оптимизированная загрузка
        try:
            print('Loading logo...')
            logo_url = 'https://cdn.poehali.dev/files/%D0%BB%D0%BE%D0%B3%D0%BE%D0%BA%D0%BF.png'
            req = urllib.request.Request(logo_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=3) as response:
                logo_data = response.read()
                
                # Открываем и сразу ресайзим для экономии памяти
                pil_logo = PILImage.open(io.BytesIO(logo_data))
                
                # Целевые размеры
                target_width = 180
                target_height = 90
                
                # Ресайзим изображение сразу
                pil_logo.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)
                
                # Сохраняем в буфер
                logo_buffer = io.BytesIO()
                pil_logo.save(logo_buffer, format='PNG', optimize=True)
                logo_buffer.seek(0)
                
                logo_img = XLImage(logo_buffer)
                logo_img.width = pil_logo.width
                logo_img.height = pil_logo.height
                
                ws.add_image(logo_img, 'B2')
                print('Logo loaded successfully')
        except Exception as e:
            print(f'Failed to load logo: {e}')
        
        ws.merge_cells(f'A{current_row}:B{current_row+4}')
        
        # Шапка компании (правый верхний угол) - начинаем с колонки C чтобы было левее
        ws.merge_cells(f'C{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=3, value='ИП ПРОНИН РУСЛАН ОЛЕГОВИЧ')
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # ИНН и ОГРНИП
        ws.merge_cells(f'C{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=3, value='ИНН 110209455200 ОГРНИП 32377460012482')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # Адрес
        ws.merge_cells(f'C{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=3, value='350005, г. Краснодар, ул. Кореновская, д. 57 оф.7')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # Телефон и email
        ws.merge_cells(f'C{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=3, value='тел: +7 918 115 15 51 e-mail: info@urban-play.ru')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # Сайт (кликабельная ссылка) - фиолетовый цвет
        ws.merge_cells(f'C{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=3)
        cell.value = '=HYPERLINK("https://urban-play.ru", "urban-play.ru")'
        cell.font = Font(name='Calibri', size=11, color='59068c', underline='single')
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # Декоративные линии (салатовая и фиолетовая) - очень тонкие
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.fill = PatternFill(start_color='44aa02', end_color='44aa02', fill_type='solid')  # Зелёный
        ws.row_dimensions[current_row].height = 2
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.fill = PatternFill(start_color='58078a', end_color='58078a', fill_type='solid')  # Фиолетовый
        ws.row_dimensions[current_row].height = 2
        current_row += 1
        
        # Пустая строка для отступа
        current_row += 1
        
        # Заголовок КП
        ws.merge_cells(f'A{current_row}:G{current_row}')
        kp_title = f'Коммерческое предложение № {kp_number:04d} от {datetime.now().strftime("%d.%m.%Y")}'
        cell = ws.cell(row=current_row, column=1, value=kp_title)
        cell.font = Font(name='Calibri', bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 2
        
        # Адрес объекта (если указан)
        if address:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws.cell(row=current_row, column=1, value=f'Адрес объекта: {address}')
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[current_row].height = 15
            current_row += 2
        
        # Настройка колонок
        ws.column_dimensions['A'].width = 4.00   # № - 4.00
        ws.column_dimensions['B'].width = 27.00  # Наименование - 27.00
        ws.column_dimensions['C'].width = 20.00  # Рисунок - 20.00
        ws.column_dimensions['D'].width = 7.00   # Кол-во - 7.00
        ws.column_dimensions['E'].width = 7.00   # Ед. изм - 7.00
        ws.column_dimensions['F'].width = 13.00  # Цена руб - 13.00
        ws.column_dimensions['G'].width = 14.00  # Сумма руб - 14.00
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки таблицы - светло-фиолетовый фон
        headers = ['№', 'Наименование', 'Рисунок', 'Кол-во', 'Ед. изм', 'Цена, руб', 'Сумма, руб']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(name='Calibri', bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color='e8d9f0', end_color='e8d9f0', fill_type='solid')
            cell.font = Font(name='Calibri', bold=True, size=10, color='000000')
        
        ws.row_dimensions[current_row].height = 16.50
        current_row += 1
        
        # Товары
        equipment_total_original = 0  # Сумма товаров без скидки
        
        # Рассчитываем общее количество товаров для равномерного распределения доставки
        total_product_quantity = sum(p['quantity'] for p in products)
        
        # Сначала считаем оригинальную сумму товаров для расчета скидки
        for product in products:
            base_price = int(product['price'].replace(' ', '')) if isinstance(product['price'], str) else product['price']
            equipment_total_original += base_price * product['quantity']
        
        # Рассчитываем скидку от суммы товаров
        discount_value = 0
        if discount_amount > 0:
            discount_value = discount_amount
        elif discount_percent > 0:
            discount_value = equipment_total_original * (discount_percent / 100)
        
        # Монтаж считается от суммы товаров ДО скидки
        calculated_installation_cost = equipment_total_original * (installation_percent / 100) if installation_percent > 0 else 0
        
        # Доставка на единицу товара (равномерное распределение, если скрыта)
        delivery_per_unit = (delivery_cost / total_product_quantity) if (hide_delivery and delivery_cost > 0 and total_product_quantity > 0) else 0
        
        # Процент монтажа для добавления к цене (если скрыт)
        installation_per_unit = (calculated_installation_cost / total_product_quantity) if (hide_installation and calculated_installation_cost > 0 and total_product_quantity > 0) else 0
        
        equipment_total = 0  # Сумма товаров (полная, БЕЗ скидки)
        
        for idx, product in enumerate(products, 1):
            ws.row_dimensions[current_row].height = 75.00
            
            # №
            cell = ws.cell(row=current_row, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            # Наименование (сначала название, потом артикул)
            article = product.get('article', '')
            name = product.get('name', '')
            full_name = f"{name}\n{article}" if article else name
            cell = ws.cell(row=current_row, column=2, value=full_name)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            # Рисунок - оптимизированная загрузка
            if product.get('image') and product['image'].startswith('http'):
                try:
                    print(f'Loading product image {idx}...')
                    # Энкодим URL для корректной работы с кириллицей
                    image_url = product['image']
                    
                    # Парсим URL
                    parsed = urllib.parse.urlparse(image_url)
                    encoded_path = urllib.parse.quote(parsed.path, safe='/')
                    safe_url = urllib.parse.urlunparse((
                        parsed.scheme, parsed.netloc, encoded_path,
                        parsed.params, parsed.query, parsed.fragment
                    ))
                    
                    req = urllib.request.Request(safe_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req, timeout=3) as response:
                        img_data = response.read()
                        
                        # Открываем и сразу ресайзим
                        pil_img = PILImage.open(io.BytesIO(img_data))
                        
                        # Целевые размеры
                        target_width = 130
                        target_height = 90
                        
                        # Ресайзим сразу для экономии памяти
                        pil_img.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)
                        
                        # Сохраняем в буфер
                        img_buffer = io.BytesIO()
                        pil_img.save(img_buffer, format='PNG', optimize=True)
                        img_buffer.seek(0)
                        
                        img = XLImage(img_buffer)
                        img.width = pil_img.width
                        img.height = pil_img.height
                        
                        # Центрируем изображение (EMU: 1 px ≈ 9525 EMU)
                        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
                        
                        col_width_pixels = 140
                        row_height_pixels = 100
                        
                        # Рассчитываем отступы для центрирования
                        offset_x = int((col_width_pixels - pil_img.width) / 2 * 9525)
                        offset_y = int((row_height_pixels - pil_img.height) / 2 * 9525)
                        
                        # Привязка к колонке C (индекс 2)
                        anchor = TwoCellAnchor()
                        anchor._from = AnchorMarker(col=2, colOff=offset_x, row=current_row-1, rowOff=offset_y)
                        anchor.to = AnchorMarker(col=2, colOff=offset_x + pil_img.width * 9525, row=current_row-1, rowOff=offset_y + pil_img.height * 9525)
                        img.anchor = anchor
                        
                        ws.add_image(img)
                        print(f'Product image {idx} loaded')
                except Exception as e:
                    print(f'Failed to load image {idx}: {e}')
            
            cell = ws.cell(row=current_row, column=3, value='')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Кол-во
            quantity = product['quantity']
            cell = ws.cell(row=current_row, column=4, value=quantity)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            # Ед. изм
            cell = ws.cell(row=current_row, column=5, value='шт')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            # Цена ПОЛНАЯ (БЕЗ скидки)
            base_price = int(product['price'].replace(' ', '')) if isinstance(product['price'], str) else product['price']
            
            # Проверяем артикул: для Благоустройства (9000-9050) монтаж НЕ применяется
            article = product.get('article', '')
            exclude_installation = False
            try:
                article_num = int(article) if article.isdigit() else 0
                if 9000 <= article_num <= 9050:
                    exclude_installation = True
            except:
                pass
            
            # Если монтаж скрыт - добавляем его равномерно на каждый товар (кроме исключений)
            if exclude_installation:
                price_with_installation = base_price
            else:
                price_with_installation = base_price + (installation_per_unit / quantity if quantity > 0 else 0)
            
            # Если доставка скрыта - добавляем равномерно на единицу товара
            final_price = price_with_installation + delivery_per_unit
            
            cell = ws.cell(row=current_row, column=6, value=final_price)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            # Сумма (ПОЛНАЯ, без скидки)
            final_sum = final_price * quantity
            equipment_total += final_sum
            cell = ws.cell(row=current_row, column=7, value=final_sum)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            current_row += 1
        
        # Монтаж (если не скрыт) - считаем от суммы товаров со скидкой
        if calculated_installation_cost > 0 and not hide_installation:
            ws.row_dimensions[current_row].height = 25
            
            cell = ws.cell(row=current_row, column=1, value=len(products) + 1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=2, value=f'Монтаж ({installation_percent}%)')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=3, value='')
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=4, value=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=5, value='усл')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=6, value=calculated_installation_cost)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=7, value=calculated_installation_cost)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            current_row += 1
        
        # Доставка (если не скрыта)
        if delivery_cost > 0 and not hide_delivery:
            ws.row_dimensions[current_row].height = 25
            
            next_num = len(products) + (2 if (installation_cost > 0 and not hide_installation) else 1)
            cell = ws.cell(row=current_row, column=1, value=next_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=2, value='Доставка')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=3, value='')
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=4, value=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=5, value='усл')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=6, value=delivery_cost)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            cell = ws.cell(row=current_row, column=7, value=delivery_cost)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            
            current_row += 1
        
        # Итого = товары (без скидки) + монтаж (от суммы со скидкой) + доставка
        total_before_discount = equipment_total_original
        if calculated_installation_cost > 0 and not hide_installation:
            total_before_discount += calculated_installation_cost
        if delivery_cost > 0 and not hide_delivery:
            total_before_discount += delivery_cost
        
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col, value='')
        
        cell = ws.cell(row=current_row, column=6, value='Итого:')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=7, value=total_before_discount)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '#,##0.00\ ""'
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.border = thin_border
        current_row += 1
        
        # Скидка (уже рассчитана выше)
        if discount_value > 0:
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col, value='')
            
            cell = ws.cell(row=current_row, column=6, value='Скидка:')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, color='FF0000')
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=7, value=-abs(discount_value))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0.00\ ""'
            cell.font = Font(name='Calibri', size=11, color='FF0000')
            cell.border = thin_border
            current_row += 1
        
        # К оплате = Итого - Скидка (должно равняться целевой сумме)
        total_to_pay = total_before_discount - discount_value
        
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col, value='')
        
        cell = ws.cell(row=current_row, column=6, value='К оплате:')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=7, value=total_to_pay)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '#,##0.00\ ""'
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.border = thin_border
        
        current_row += 1
        
        # Футер с условиями
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1, value='Вся продукция сертифицирована и соответствует стандартам качества')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1, value='Срок действия коммерческого предложения 15 дней')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1, value='Срок изготовления оборудования 30 дней')
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 2
        
        # Подпись
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1, value='Индивидуальный предприниматель___________________________/Пронин Р.О./')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=11)
        
        # Сохранение
        print('Saving Excel file...')
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.read()
        print(f'Excel file saved, size: {len(excel_data)} bytes')
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename="commercial_offer.xlsx"',
                'Access-Control-Allow-Origin': '*'
            },
            'body': base64.b64encode(excel_data).decode('utf-8'),
            'isBase64Encoded': True
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f'ERROR in handler: {error_details}')
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': str(e), 'details': error_details})
        }