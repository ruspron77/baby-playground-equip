import json
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
import requests

def handler(event: dict, context) -> dict:
    '''Генерация Excel-шаблона для загрузки каталога с примерами'''
    
    method = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': ''
        }
    
    if method == 'GET':
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог товаров"
        
        # Заголовки
        headers = ['Картинка', 'Категория', 'Артикул', 'Название', 'Размеры', 'Цена']
        ws.append(headers)
        
        # Стилизация заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Примеры данных
        examples = [
            ['', 'Столы', 'ST-001', 'Стол обеденный "Классик"', '120x80x75', '15990'],
            ['', 'Стулья', 'CH-001', 'Стул "Комфорт"', '45x50x95', '4990'],
            ['', 'Шкафы', 'WR-001', 'Шкаф-купе "Престиж"', '200x240x60', '35990']
        ]
        
        for row_data in examples:
            ws.append(row_data)
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 20  # Картинка
        ws.column_dimensions['B'].width = 15  # Категория
        ws.column_dimensions['C'].width = 12  # Артикул
        ws.column_dimensions['D'].width = 30  # Название
        ws.column_dimensions['E'].width = 15  # Размеры
        ws.column_dimensions['F'].width = 10  # Цена
        
        # Высота строк
        ws.row_dimensions[1].height = 25
        for i in range(2, 5):
            ws.row_dimensions[i].height = 80
        
        # Добавляем примечание
        ws['A6'] = 'Инструкция:'
        ws['A7'] = '1. Вставьте изображения товаров в столбец "Картинка" (ПКМ → Вставить → Рисунок)'
        ws['A8'] = '2. Заполните данные по каждому товару'
        ws['A9'] = '3. Удалите примеры (строки 2-4)'
        ws['A10'] = '4. Сохраните и загрузите файл в админ-панель'
        
        ws['A6'].font = Font(bold=True, size=11)
        
        # Сохранение в BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Конвертация в base64
        excel_base64 = base64.b64encode(excel_file.read()).decode('utf-8')
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'file': excel_base64,
                'filename': 'shablon_kataloga.xlsx'
            })
        }
    
    return {
        'statusCode': 405,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({'error': 'Method not allowed'})
    }
