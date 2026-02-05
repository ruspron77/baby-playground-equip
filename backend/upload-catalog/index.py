import json
import base64
import os
import boto3
import openpyxl
import re
import psycopg2
from io import BytesIO
import uuid

def handler(event: dict, context) -> dict:
    '''Загрузка Excel-файла с каталогом, извлечение изображений и сохранение в базу данных'''
    
    method = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        filename = body.get('filename', 'catalog.xlsx')
        base64_content = body.get('content', '')
        update_mode = body.get('updateMode', 'new')
        
        if not base64_content:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({
                    'success': False,
                    'error': 'Файл не передан'
                }),
                'isBase64Encoded': False
            }
        
        file_data = base64.b64decode(base64_content)
        
        # Инициализируем S3
        s3 = boto3.client('s3',
            endpoint_url='https://bucket.poehali.dev',
            aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
            aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY']
        )
        
        # Подключаемся к БД
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cursor = conn.cursor()
        schema = os.environ['MAIN_DB_SCHEMA']
        
        # Парсим Excel
        workbook = openpyxl.load_workbook(BytesIO(file_data))
        products_count = 0
        images_uploaded = 0
        updated_count = 0
        added_count = 0
        
        # Словарь для сопоставления позиций изображений с артикулами
        image_map = {}
        
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            
            if 'оглавление' in sheet_name.lower():
                continue
            
            # Извлекаем изображения из листа
            print(f'Sheet {sheet_name}: checking for images, has _images={hasattr(sheet, "_images")}')
            
            if hasattr(sheet, '_images'):
                print(f'Sheet {sheet_name}: found {len(sheet._images)} images')
                
                for idx, image in enumerate(sheet._images):
                    try:
                        img_data = image._data()
                        img_ext = image.format.lower() if hasattr(image, 'format') else 'png'
                        img_filename = f"catalog-images/{uuid.uuid4()}.{img_ext}"
                        
                        # Загружаем изображение в S3
                        s3.put_object(
                            Bucket='files',
                            Key=img_filename,
                            Body=img_data,
                            ContentType=f'image/{img_ext}'
                        )
                        
                        # Сохраняем URL и позицию
                        img_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{img_filename}"
                        
                        # Получаем позицию изображения (строка)
                        anchor = image.anchor
                        if hasattr(anchor, '_from'):
                            row_idx = anchor._from.row
                            print(f'Image {idx}: uploaded to row {row_idx}, url={img_url[:60]}...')
                            image_map[row_idx] = img_url
                        else:
                            print(f'Image {idx}: no anchor._from, anchor type={type(anchor)}')
                        
                        images_uploaded += 1
                    except Exception as img_error:
                        print(f'Error uploading image {idx}: {img_error}')
            else:
                print(f'Sheet {sheet_name}: NO _images attribute found!')
            
            # Ищем заголовки
            header_row_idx = None
            col_indices = {}
            
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=False), start=1):
                row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
                
                if any('артикул' in val for val in row_values):
                    header_row_idx = row_idx
                    
                    # Определяем индексы колонок
                    for col_idx, val in enumerate(row_values):
                        if 'картинк' in val or 'фото' in val or 'изображен' in val:
                            col_indices['image'] = col_idx
                        elif 'подподподподкатегория' in val:
                            col_indices['subcategory4'] = col_idx
                        elif 'подподподкатегория' in val:
                            col_indices['subcategory3'] = col_idx
                        elif 'подподкатегория' in val:
                            col_indices['subcategory2'] = col_idx
                        elif 'подкатегория' in val:
                            col_indices['subcategory1'] = col_idx
                        elif 'категория' in val:
                            col_indices['category'] = col_idx
                        elif 'артикул' in val:
                            col_indices['article'] = col_idx
                        elif 'название' in val or 'наименование' in val:
                            col_indices['name'] = col_idx
                        elif 'габарит' in val or 'размер' in val:
                            col_indices['dimensions'] = col_idx
                        elif 'ед.' in val and 'изм' in val:
                            col_indices['unit'] = col_idx
                        elif 'цена' in val:
                            col_indices['price'] = col_idx
                    break
            
            if header_row_idx is None:
                print(f'Sheet {sheet_name}: header not found')
                continue
            
            print(f'Sheet {sheet_name}: found header at row {header_row_idx}, columns: {col_indices}')
            
            # Читаем товары
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                if not any(str(val).strip() if val else '' for val in row):
                    continue
                
                # Извлекаем данные: Картинка(0), Категория(1), Подкатегория(2), Подподкатегория(3), Подподподкатегория(4), Подподподподкатегория(5), Артикул, Название, Размеры, Цена
                category = str(row[col_indices.get('category', 1)]).strip() if col_indices.get('category') is not None and len(row) > col_indices.get('category', 1) and row[col_indices.get('category', 1)] else 'Без категории'
                subcategory1 = str(row[col_indices.get('subcategory1', 2)]).strip() if col_indices.get('subcategory1') is not None and len(row) > col_indices.get('subcategory1', 2) and row[col_indices.get('subcategory1', 2)] else ''
                subcategory2 = str(row[col_indices.get('subcategory2', 3)]).strip() if col_indices.get('subcategory2') is not None and len(row) > col_indices.get('subcategory2', 3) and row[col_indices.get('subcategory2', 3)] else ''
                subcategory3 = str(row[col_indices.get('subcategory3', 4)]).strip() if col_indices.get('subcategory3') is not None and len(row) > col_indices.get('subcategory3', 4) and row[col_indices.get('subcategory3', 4)] else ''
                subcategory4 = str(row[col_indices.get('subcategory4', 5)]).strip() if col_indices.get('subcategory4') is not None and len(row) > col_indices.get('subcategory4', 5) and row[col_indices.get('subcategory4', 5)] else ''
                
                # Собираем полный путь категории
                category_parts = [category]
                if subcategory1 and subcategory1 != 'None':
                    category_parts.append(subcategory1)
                if subcategory2 and subcategory2 != 'None':
                    category_parts.append(subcategory2)
                if subcategory3 and subcategory3 != 'None':
                    category_parts.append(subcategory3)
                if subcategory4 and subcategory4 != 'None':
                    category_parts.append(subcategory4)
                full_category = ' > '.join(category_parts)
                
                article = str(row[col_indices.get('article', 4)]).strip() if col_indices.get('article') is not None and len(row) > col_indices.get('article', 4) and row[col_indices.get('article', 4)] else ''
                name = str(row[col_indices.get('name', 5)]).strip() if col_indices.get('name') is not None and len(row) > col_indices.get('name', 5) and row[col_indices.get('name', 5)] else ''
                dimensions = str(row[col_indices.get('dimensions', 6)]).strip() if col_indices.get('dimensions') is not None and len(row) > col_indices.get('dimensions', 6) and row[col_indices.get('dimensions', 6)] else ''
                unit = str(row[col_indices.get('unit', 8)]).strip() if col_indices.get('unit') is not None and len(row) > col_indices.get('unit', 8) and row[col_indices.get('unit', 8)] else 'шт'
                
                # Обработка цены
                price_val = row[col_indices.get('price', 9)] if col_indices.get('price') is not None and len(row) > col_indices.get('price', 9) else 0
                try:
                    price = int(float(str(price_val).replace(' ', '').replace(',', '.').replace('₽', '').strip())) if price_val else 0
                except:
                    price = 0
                
                if not article and not name:
                    continue
                
                # Ищем изображение для этой строки
                image_url = image_map.get(row_idx, None)
                
                print(f'Processing: article={article}, name={name[:30] if len(name) > 30 else name}, category={full_category}, image={bool(image_url)}')
                
                # Escape для Simple Query Protocol
                safe_article = article.replace("'", "''")
                safe_name = name.replace("'", "''")
                safe_category = full_category.replace("'", "''")
                safe_dimensions = dimensions.replace("'", "''")
                safe_image_url = (image_url or '').replace("'", "''")
                safe_unit = unit.replace("'", "''")
                
                # Проверяем существует ли товар
                cursor.execute(f"SELECT id FROM {schema}.products WHERE article = '{safe_article}'")
                existing = cursor.fetchone()
                
                if update_mode == 'update':
                    # Режим обновления - обновляем существующие или добавляем новые
                    if existing:
                        cursor.execute(f"""
                            UPDATE {schema}.products 
                            SET name = '{safe_name}',
                                category = '{safe_category}',
                                price = {price},
                                dimensions = '{safe_dimensions}',
                                unit = '{safe_unit}',
                                image_url = CASE WHEN '{safe_image_url}' != '' THEN '{safe_image_url}' ELSE image_url END
                            WHERE article = '{safe_article}'
                        """)
                        updated_count += 1
                    else:
                        cursor.execute(f"""
                            INSERT INTO {schema}.products (article, name, category, price, dimensions, unit, image_url)
                            VALUES ('{safe_article}', '{safe_name}', '{safe_category}', {price}, '{safe_dimensions}', '{safe_unit}', '{safe_image_url}')
                        """)
                        added_count += 1
                else:
                    # Режим добавления - только новые товары
                    if not existing:
                        cursor.execute(f"""
                            INSERT INTO {schema}.products (article, name, category, price, dimensions, unit, image_url)
                            VALUES ('{safe_article}', '{safe_name}', '{safe_category}', {price}, '{safe_dimensions}', '{safe_unit}', '{safe_image_url}')
                        """)
                        added_count += 1
                
                products_count += 1
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'success': True,
                'productsCount': products_count,
                'imagesUploaded': images_uploaded,
                'updatedCount': updated_count,
                'addedCount': added_count,
                'message': f'Файл загружен успешно. Обработано товаров: {products_count}, изображений: {images_uploaded}, обновлено: {updated_count}, добавлено: {added_count}'
            }, ensure_ascii=False),
            'isBase64Encoded': False
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f'Error: {error_details}')
        
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'success': False,
                'error': str(e)
            }),
            'isBase64Encoded': False
        }